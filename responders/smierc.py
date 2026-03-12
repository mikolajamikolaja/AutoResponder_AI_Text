"""
responders/smierc.py
Pośmiertny autoresponder — uniwersalny, sterowany przez CSV.

Plik konfiguracyjny: prompts/requiem_etapy.csv
Kolumny CSV:
  etap          — numer etapu (int)
  opis          — treść/myśl przewodnia etapu
  obraz         — lista plików z media/images/niebo/ oddzielona przecinkami (opcjonalna)
  video         — lista plików z media/mp4/niebo/ oddzielona przecinkami (opcjonalna)
  system_prompt — pełny system prompt dla AI (opcjonalny — fallback do DEFAULT_SYSTEM_PROMPT)
  styl_flux     — styl obrazka FLUX:
                  * nazwa pliku .txt → czyta plik z prompts/
                  * inny tekst       → używa wprost
                  * puste            → brak stylu

Logika obrazka:
  obraz wpisany  → wyślij statyczne pliki z media/images/niebo/
  obraz pusty    → generuj przez FLUX:
      1. Groq (flux_groq_system.txt) generuje prompt
      2. DeepSeek fallback — ten sam system prompt
      3. Oba padły → tekst nadawcy + styl z CSV (jeśli jest)
  brak tokenów HF → wyślij tylko wiadomość + _.txt z promptem

Logika etapów:
  etap w CSV     → obsłuż normalnie
  etap > max CSV → tryb Wysłannika (DeepSeek + FLUX)
"""

import os
import re
import random
import base64
import requests
from concurrent.futures import ThreadPoolExecutor, as_completed
from flask import current_app

from core.ai_client import call_deepseek, MODEL_TYLER

BASE_DIR    = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
PROMPTS_DIR = os.path.join(BASE_DIR, "prompts")
MEDIA_DIR   = os.path.join(BASE_DIR, "media")

# ── Ścieżki plików ────────────────────────────────────────────────────────────
FILE_XLSX                    = os.path.join(PROMPTS_DIR, "requiem_etapy.xlsx")
FILE_WYSLANNIK_SYSTEM        = os.path.join(PROMPTS_DIR, "requiem_WYSLANNIK_system_8_.txt")
FILE_WYSLANNIK_FLUX_GROQ_SYS = os.path.join(PROMPTS_DIR, "requiem_WYSLANNIK_flux_groq_system.txt")
FILE_FLUX_FORBIDDEN          = os.path.join(PROMPTS_DIR, "flux_forbidden.txt")
FILE_FLUX_MUTATIONS          = os.path.join(PROMPTS_DIR, "flux_mutations.txt")

# ── Stałe FLUX ────────────────────────────────────────────────────────────────
HF_API_URL  = "https://router.huggingface.co/hf-inference/models/black-forest-labs/FLUX.1-schnell"
HF_STEPS    = 5
HF_GUIDANCE = 5
TIMEOUT_SEC = 55

# ── Groq ──────────────────────────────────────────────────────────────────────
GROQ_MODEL   = "llama-3.3-70b-versatile"
GROQ_API_URL = "https://api.groq.com/openai/v1/chat/completions"

# ── Limit obrazków AI (zabezpieczenie przed za dużą liczbą) ──────────────────
MAX_AI_IMAGES = 10
DEFAULT_SYSTEM_PROMPT = (
    "Piszesz tajemniczą wiadomość z innego wymiaru. "
    "Ton: poetycki, absurdalny, ciepły. Maksymalnie 5 zdań. "
    "Nawiąż do wiadomości nadawcy."
)


# ═══════════════════════════════════════════════════════════════════════════════
# NARZĘDZIA
# ═══════════════════════════════════════════════════════════════════════════════

def _load_txt(path: str, fallback: str = "") -> str:
    try:
        with open(path, encoding="utf-8") as f:
            return f.read().strip()
    except Exception as e:
        current_app.logger.warning("[txt] Błąd wczytywania %s: %s", path, e)
        return fallback


def _resolve_styl_flux(styl_raw: str) -> str:
    """
    Jeśli styl_raw kończy się na .txt → wczytaj plik z prompts/.
    W przeciwnym razie użyj tekstu wprost.
    Zwraca pusty string gdy styl_raw pusty.
    """
    if not styl_raw:
        return ""
    if styl_raw.lower().endswith(".txt"):
        path = os.path.join(PROMPTS_DIR, styl_raw)
        content = _load_txt(path, fallback="")
        if content:
            current_app.logger.info("[styl] Wczytano plik stylu: %s", styl_raw)
            return content
        current_app.logger.warning("[styl] Brak pliku stylu: %s — używam nazwy jako tekst", path)
        return styl_raw  # fallback: użyj nazwy pliku wprost
    return styl_raw


# ── XLSX ──────────────────────────────────────────────────────────────────────

def _parse_int(val) -> int | None:
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return None


def _load_xlsx() -> dict:
    """
    Wczytuje requiem_etapy.xlsx — dwie zakładki:
      'etapy' : etap | opis | obraz | video | system_prompt
      'style'  : etap | styl

    Łączy po numerze etapu.
    Zwraca {etap_int: {opis, obraz, video, system_prompt, styl_flux}}.
    """
    import openpyxl
    etapy  = {}
    style_map = {}

    try:
        wb = openpyxl.load_workbook(FILE_XLSX, read_only=True, data_only=True)

        # ── zakładka style ────────────────────────────────────────────────────
        if "style" in wb.sheetnames:
            ws_s = wb["style"]
            rows_s = list(ws_s.iter_rows(values_only=True))
            # nagłówek w wierszu 1 — pomijamy, kolumny: A=etap B=styl
            for row in rows_s[1:]:
                etap_num = _parse_int(row[0] if row else None)
                styl_val = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ""
                if etap_num is not None:
                    style_map[etap_num] = styl_val

        # ── zakładka etapy ────────────────────────────────────────────────────
        ws_e  = wb["etapy"]
        rows_e = list(ws_e.iter_rows(values_only=True))
        if not rows_e:
            wb.close()
            return etapy

        headers = [str(h).strip().lower() if h else "" for h in rows_e[0]]
        for row in rows_e[1:]:
            row_dict = {headers[i]: (str(v).strip() if v is not None else "")
                        for i, v in enumerate(row) if i < len(headers)}
            etap_num = _parse_int(row_dict.get("etap"))
            if etap_num is None:
                continue
            # Szukaj obrazki_ai tolerując spacje w nazwie nagłówka
            obrazki_ai_val = (
                row_dict.get("obrazki_ai")
                or row_dict.get("obrazki_ai ")
                or "0"
            )
            etapy[etap_num] = {
                "opis":          row_dict.get("opis",          ""),
                "obraz":         row_dict.get("obraz",         ""),
                "video":         row_dict.get("video",         ""),
                "obrazki_ai":    _parse_int(obrazki_ai_val) or 0,
                "system_prompt": row_dict.get("system_prompt", ""),
                "styl_flux":     style_map.get(etap_num, ""),
            }

        wb.close()

    except Exception as e:
        current_app.logger.warning("[xlsx] Błąd wczytywania %s: %s", FILE_XLSX, e)

    return etapy


# ── Media ─────────────────────────────────────────────────────────────────────

def _file_to_base64(path: str):
    try:
        with open(path, "rb") as f:
            return base64.b64encode(f.read()).decode("ascii")
    except Exception:
        return None


def _guess_content_type(filename: str) -> str:
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else ""
    return {
        "png":  "image/png",
        "jpg":  "image/jpeg",
        "jpeg": "image/jpeg",
        "bmp":  "image/bmp",
        "gif":  "image/gif",
        "webp": "image/webp",
        "mp4":  "video/mp4",
        "mov":  "video/quicktime",
        "avi":  "video/avi",
        "pdf":  "application/pdf",
        "doc":  "application/msword",
        "docx": "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    }.get(ext, "application/octet-stream")


def _compress_to_jpg(png_base64: str, quality: int) -> str:
    """
    Konwertuje PNG (base64) → JPG (base64) z podaną jakością.
    Zwraca base64 JPG lub oryginalny PNG base64 gdy konwersja się nie uda.
    """
    try:
        from PIL import Image
        import io
        png_bytes = base64.b64decode(png_base64)
        img = Image.open(io.BytesIO(png_bytes)).convert("RGB")
        buf = io.BytesIO()
        img.save(buf, format="JPEG", quality=quality, optimize=True)
        return base64.b64encode(buf.getvalue()).decode("ascii")
    except Exception as e:
        current_app.logger.warning("[compress] Błąd kompresji JPG: %s", e)
        return png_base64  # fallback: zwróć oryginał


def _jpg_quality(n_obrazkow: int) -> int:
    """Zwraca jakość JPG zależnie od liczby obrazków AI."""
    if n_obrazkow == 1: return 90
    if n_obrazkow == 2: return 80
    if n_obrazkow == 3: return 85
    if n_obrazkow == 4: return 80
    return 50  # 5 i więcej


def _compress_images_ai(images: list, n_obrazkow: int) -> list:
    """
    Kompresuje listę obrazków AI (PNG→JPG) gdy n_obrazkow >= 1.
    Zwraca listę ze zaktualizowanymi base64, content_type i filename.
    """
    if not images:
        return images
    quality = _jpg_quality(n_obrazkow)
    current_app.logger.info("[compress] %d obrazków → JPG quality=%d", len(images), quality)
    result = []
    for img in images:
        compressed_b64 = _compress_to_jpg(img["base64"], quality)
        fname = img["filename"].rsplit(".", 1)[0] + ".jpg"
        result.append({
            "base64":       compressed_b64,
            "content_type": "image/jpeg",
            "filename":     fname,
        })
    return result
    """
    Parsuje string z listą plików oddzielonych przecinkami.
    Szuka każdego pliku w base_dir.
    Zwraca listę {base64, content_type, filename}.
    """
    results = []
    if not file_list_str:
        return results
    for fname in file_list_str.split(","):
        fname = fname.strip()
        if not fname:
            continue
        path = os.path.join(base_dir, fname)
        b64  = _file_to_base64(path)
        if b64:
            results.append({
                "base64":       b64,
                "content_type": _guess_content_type(fname),
                "filename":     fname,
            })
            current_app.logger.info("[media] OK: %s", fname)
        else:
            current_app.logger.warning("[media] Brak pliku: %s", path)
    return results


def _load_images(file_list_str: str) -> list:
    return _load_file_list(file_list_str, os.path.join(MEDIA_DIR, "images", "niebo"))


def _load_videos(file_list_str: str) -> list:
    return _load_file_list(file_list_str, os.path.join(MEDIA_DIR, "mp4", "niebo"))


# ── AI ────────────────────────────────────────────────────────────────────────

def _call_groq(system: str, user: str) -> str | None:
    api_key = os.getenv("API_KEY_GROQ", "").strip()
    if not api_key:
        current_app.logger.warning("[groq] Brak API_KEY_GROQ")
        return None
    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type":  "application/json",
    }
    payload = {
        "model":       GROQ_MODEL,
        "messages":    [
            {"role": "system", "content": system},
            {"role": "user",   "content": user},
        ],
        "max_tokens":  300,
        "temperature": 0.95,
    }
    try:
        resp = requests.post(GROQ_API_URL, headers=headers, json=payload, timeout=30)
        if resp.status_code == 200:
            result = resp.json()["choices"][0]["message"]["content"].strip()
            current_app.logger.info("[groq] OK: %.150s", result)
            return result
        current_app.logger.warning("[groq] HTTP %s: %s", resp.status_code, resp.text[:150])
        return None
    except Exception as e:
        current_app.logger.warning("[groq] Wyjątek: %s", str(e)[:100])
        return None


# ── Mutacja FLUX ──────────────────────────────────────────────────────────────

def _load_word_list(path: str) -> list:
    words = []
    try:
        with open(path, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if line and not line.startswith("#"):
                    words.append(line.lower())
    except Exception as e:
        current_app.logger.warning("[wordlist] Błąd %s: %s", path, e)
    return words


def _mutate_flux_prompt(prompt: str) -> tuple:
    """
    Zamienia zabronione słowa na słowo-of-randomowy-sufiks.
    Zwraca (zmutowany_prompt, lista_zmian).
    """
    forbidden = _load_word_list(FILE_FLUX_FORBIDDEN)
    suffixes  = _load_word_list(FILE_FLUX_MUTATIONS)
    if not forbidden or not suffixes:
        return prompt, []
    result  = prompt
    changes = []
    for word in forbidden:
        pattern = re.compile(
            rf'(?<![a-zA-Z]){re.escape(word)}(?![a-zA-Z])',
            re.IGNORECASE
        )
        if pattern.search(result):
            sufiks = random.choice(suffixes)
            result = pattern.sub(lambda m, s=sufiks: m.group(0) + "-" + s, result)
            changes.append(f"{word} → {word}-{sufiks}")
    current_app.logger.info("[mutate] Zmutowano słów: %d", len(changes))
    return result, changes


# ── Generowanie promptu FLUX ──────────────────────────────────────────────────

def _generate_flux_prompt(source_text: str, styl_flux: str = "") -> tuple:
    """
    Kolejność:
      1. Groq  (flux_groq_system.txt)
      2. DeepSeek fallback (ten sam system prompt)
      3. Oba padły → source_text + styl_flux (jeśli jest)

    Zwraca (prompt_po_mutacji, lista_zmian, provider_name).
    """
    system = _load_txt(FILE_WYSLANNIK_FLUX_GROQ_SYS, fallback=(
        "You are a creative prompt engineer for FLUX image generator. "
        "Write a surreal, otherworldly image prompt in English (max 80 words). "
        "Return ONLY the prompt."
    ))
    user = f"Generate a FLUX image prompt based on this text:\n\n{source_text}"

    # 1. Groq
    result   = _call_groq(system, user)
    provider = "Groq"

    # 2. DeepSeek fallback
    if not result:
        current_app.logger.warning("[flux-prompt] Groq zawiódł — próbuję DeepSeek")
        try:
            result   = call_deepseek(system, user, MODEL_TYLER)
            provider = "DeepSeek (fallback po Groq)"
        except Exception as e:
            current_app.logger.warning("[flux-prompt] DeepSeek wyjątek: %s", e)
            result = None

    # 3. Fallback — tekst nadawcy (bez stylu żeby nie dodawać śmieci)
    if not result:
        current_app.logger.warning("[flux-prompt] Oba API zawiodły — fallback na tekst nadawcy")
        result   = source_text[:300]
        provider = "fallback: tekst nadawcy"
        # Przy fallbacku NIE doklejamy stylu — lepszy czysty prompt niż śmieciowy
        mutated, changes = _mutate_flux_prompt(result)
        return mutated, changes, provider

    # Doklejamy styl tylko gdy prompt wygenerowany przez AI
    if styl_flux:
        result = f"{result}, {styl_flux}"

    mutated, changes = _mutate_flux_prompt(result)
    return mutated, changes, provider


# ── FLUX API ──────────────────────────────────────────────────────────────────

def _get_hf_tokens() -> list:
    names = [
        "HF_TOKEN",   "HF_TOKEN1",  "HF_TOKEN2",  "HF_TOKEN3",  "HF_TOKEN4",
        "HF_TOKEN5",  "HF_TOKEN6",  "HF_TOKEN7",  "HF_TOKEN8",  "HF_TOKEN9",
        "HF_TOKEN10", "HF_TOKEN11", "HF_TOKEN12", "HF_TOKEN13", "HF_TOKEN14",
        "HF_TOKEN15", "HF_TOKEN16", "HF_TOKEN17", "HF_TOKEN18", "HF_TOKEN19",
        "HF_TOKEN20",
    ]
    return [(n, v) for n in names if (v := os.getenv(n, "").strip())]


def _generate_flux_image(prompt: str):
    """
    Generuje obrazek FLUX.
    Zwraca {base64, content_type, filename} lub None gdy wszystkie tokeny zawiodły.
    """
    tokens = _get_hf_tokens()
    if not tokens:
        current_app.logger.error("[flux] Brak tokenów HF!")
        return None
    payload = {
        "inputs": prompt,
        "parameters": {
            "num_inference_steps": HF_STEPS,
            "guidance_scale":      HF_GUIDANCE,
        },
    }
    current_app.logger.info("[flux] Prompt: %.200s", prompt)
    for name, token in tokens:
        headers = {"Authorization": f"Bearer {token}", "Accept": "image/png"}
        try:
            resp = requests.post(HF_API_URL, headers=headers,
                                 json=payload, timeout=TIMEOUT_SEC)
            if resp.status_code == 200:
                current_app.logger.info("[flux] Sukces token=%s PNG %d B",
                                        name, len(resp.content))
                return {
                    "base64":       base64.b64encode(resp.content).decode("ascii"),
                    "content_type": "image/png",
                    "filename":     "niebo_ai.png",
                }
            elif resp.status_code in (401, 403):
                current_app.logger.warning("[flux] token %s nieważny", name)
            elif resp.status_code in (503, 529):
                current_app.logger.warning("[flux] token %s przeciążony", name)
            else:
                current_app.logger.warning("[flux] token %s błąd %s: %s",
                                           name, resp.status_code, resp.text[:100])
        except requests.exceptions.Timeout:
            current_app.logger.warning("[flux] token %s timeout", name)
        except Exception as e:
            current_app.logger.warning("[flux] token %s wyjątek: %s", name, str(e)[:50])
    current_app.logger.error("[flux] Wszystkie tokeny HF zawiodły!")
    return None


def _generate_n_flux_images(n: int, source_text: str, styl_flux: str,
                            etap: int) -> tuple:
    """
    Generuje N obrazków FLUX równolegle.
    Każdy obrazek ma INNY prompt — Groq/DeepSeek wywoływany N razy.
    Zwraca (lista_obrazków, lista_debug_info).
    lista_obrazków = [{base64, content_type, filename}, ...]  — tylko te które się udały
    lista_debug_info = lista stringów z promptami (do _.txt)
    """
    n = max(0, min(n, MAX_AI_IMAGES))
    if n == 0:
        return [], []

    current_app.logger.info("[flux-n] etap=%d generuję %d obrazków równolegle", etap, n)

    # Każde zadanie: wygeneruj prompt + obrazek
    def _single_task(idx: int):
        prompt, changes, provider = _generate_flux_prompt(source_text, styl_flux)
        image = _generate_flux_image(prompt)
        if image:
            image["filename"] = f"niebo_ai_{idx+1}.png"
        debug_line = f"[{idx+1}/{n}] provider={provider} prompt={prompt[:120]}"
        return image, debug_line, changes

    images      = []
    debug_lines = []

    with ThreadPoolExecutor(max_workers=n) as executor:
        futures = {executor.submit(_single_task, i): i for i in range(n)}
        for future in as_completed(futures):
            try:
                image, debug_line, _ = future.result()
                debug_lines.append(debug_line)
                if image:
                    images.append((futures[future], image))  # (idx, image)
            except Exception as e:
                current_app.logger.warning("[flux-n] wyjątek: %s", e)

    # Sortuj po indeksie żeby kolejność była deterministyczna
    images = [img for _, img in sorted(images, key=lambda x: x[0])]
    current_app.logger.info("[flux-n] etap=%d udanych=%d/%d", etap, len(images), n)
    return images, debug_lines


# ── Debug TXT ─────────────────────────────────────────────────────────────────

def _build_debug_txt(source_text: str, flux_info: list, etap: int) -> dict:
    """
    flux_info = lista stringów opisujących każdy wygenerowany prompt.
    """
    prompts_str = "\n".join(flux_info) if flux_info else "(brak)"
    content = (
        f"Etap: {etap}\n\n"
        f"=== PROMPTY WYSŁANE DO FLUX ===\n"
        f"{prompts_str}\n\n"
        f"=== DEBUG ===\n"
        f"Źródło: {source_text[:300]}\n\n"
        f"--- Parametry FLUX ---\n"
        f"Model: FLUX.1-schnell\n"
        f"num_inference_steps: {HF_STEPS}\n"
        f"guidance_scale: {HF_GUIDANCE}\n"
        f"API URL: {HF_API_URL}\n"
    )
    return {
        "base64":       base64.b64encode(content.encode("utf-8")).decode("ascii"),
        "content_type": "text/plain",
        "filename":     "_.txt",
    }


def _format_historia(historia: list) -> str:
    if not historia:
        return "(brak poprzednich wiadomości)"
    lines = []
    for h in historia[-3:]:
        lines.append(f"Osoba: {h.get('od', '')[:300]}")
        lines.append(f"Odpowiedź: {h.get('odpowiedz', '')[:300]}")
    return "\n".join(lines)


def _oblicz_dni(data_smierci_str: str) -> str:
    """
    Oblicza ile dni minęło od daty śmierci do dziś.
    Obsługuje formaty: YYYY-MM-DD, DD.MM.YYYY, DD/MM/YYYY, MM/DD/YYYY
    Zwraca string np. "7" lub "?" gdy nie udało się sparsować.
    """
    from datetime import date
    import re
    s = data_smierci_str.strip()
    fmt_candidates = [
        "%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%m/%d/%Y",
        "%Y/%m/%d", "%d-%m-%Y",
    ]
    # GAS często wysyła format JS: "Thu Mar 05 2026 00:00:00 GMT+0100 ..."
    # Wytnij samą datę jeśli taki format
    m = re.match(r'\w+ (\w+) (\d+) (\d{4})', s)
    if m:
        months = {"Jan":1,"Feb":2,"Mar":3,"Apr":4,"May":5,"Jun":6,
                  "Jul":7,"Aug":8,"Sep":9,"Oct":10,"Nov":11,"Dec":12}
        try:
            mn = months.get(m.group(1), 0)
            d  = date(int(m.group(3)), mn, int(m.group(2)))
            return str((date.today() - d).days)
        except Exception:
            pass
    for fmt in fmt_candidates:
        try:
            from datetime import datetime
            d = datetime.strptime(s, fmt).date()
            return str((date.today() - d).days)
        except ValueError:
            continue
    current_app.logger.warning("[dni] Nie udało się sparsować daty: %s", s)
    return "?"
    if not historia:
        return "(brak poprzednich wiadomości)"
    lines = []
    for h in historia[-3:]:
        lines.append(f"Osoba: {h.get('od', '')[:300]}")
        lines.append(f"Odpowiedź: {h.get('odpowiedz', '')[:300]}")
    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════════════════════
# TRYB WYSŁANNIKA (etap > max CSV)
# ═══════════════════════════════════════════════════════════════════════════════

def _run_wyslannik(body: str, historia: list, etap: int) -> dict:
    historia_txt = _format_historia(historia)
    system = _load_txt(FILE_WYSLANNIK_SYSTEM, fallback=(
        "Jesteś wysłannikiem z wyższych sfer duchowych piszącym po polsku. "
        "Przebijasz każdą rzecz wymienioną przez nadawcę — tylko przymiotnikami, "
        "nigdy liczbami. Ton: dostojny, poetycki, absurdalny. Max 4 zdania. "
        "Podpisz się: — Wysłannik z wyższych sfer"
    ))
    user_msg    = f"Osoba pyta: {body}\n\nHistoria:\n{historia_txt}"
    wynik_tekst = call_deepseek(system, user_msg, MODEL_TYLER)

    if not wynik_tekst:
        current_app.logger.warning("[wyslannik] DeepSeek zawiódł — próbuję Groq")
        wynik_tekst = _call_groq(system, user_msg)

    reply_html = (
        f"<p>{wynik_tekst}</p><p><i>— Wysłannik z wyższych sfer</i></p>"
        if wynik_tekst
        else "<p>Jesteśmy tu do dyspozycji.<br><i>— Wysłannik z wyższych sfer</i></p>"
    )

    # Wysłannik nie ma stylu w CSV — brak stylu to brak stylu
    flux_prompt, flux_changes, flux_provider = _generate_flux_prompt(
        wynik_tekst or body, styl_flux=""
    )
    image     = _generate_flux_image(flux_prompt)
    debug_txt = _build_debug_txt(
        wynik_tekst or "",
        [f"[1/1] provider={flux_provider} prompt={flux_prompt[:120]}"],
        etap
    )

    if not image:
        current_app.logger.warning("[wyslannik] FLUX zawiódł — tylko tekst + _.txt")

    current_app.logger.info("[wyslannik] etap=%d image=%s", etap, bool(image))
    return {
        "reply_html": reply_html,
        "nowy_etap":  etap,          # Wysłannik nie inkrementuje — zostaje na etapie
        "images":     [image] if image else [],
        "videos":     [],
        "debug_txt":  debug_txt,
    }


# ═══════════════════════════════════════════════════════════════════════════════
# GŁÓWNA FUNKCJA
# ═══════════════════════════════════════════════════════════════════════════════

def build_smierc_section(
    sender_email:     str,
    body:             str,
    etap:             int,
    data_smierci_str: str,
    historia:         list,
) -> dict:
    """
    Zwraca:
      {
        "reply_html": str,
        "nowy_etap":  int,
        "images":     [ {base64, content_type, filename}, ... ],
        "videos":     [ {base64, content_type, filename}, ... ],
        "debug_txt":  {base64, content_type, filename} | None,
      }
    """
    etapy    = _load_xlsx()
    max_etap = max(etapy.keys()) if etapy else 0

    current_app.logger.info(
        "[smierc] sender=%s etap=%d max_etap=%d",
        sender_email or "(brak)", etap, max_etap
    )

    # ── TRYB WYSŁANNIKA — etap przekroczył CSV ────────────────────────────────
    if not etapy or etap > max_etap:
        current_app.logger.info("[smierc] etap=%d > max=%d → Wysłannik", etap, max_etap)
        return _run_wyslannik(body, historia, etap)

    # ── ETAP Z XLSX ───────────────────────────────────────────────────────────
    row           = etapy[etap]
    opis          = row["opis"]
    obraz_lista   = row["obraz"]
    video_lista   = row["video"]
    obrazki_ai    = min(int(row.get("obrazki_ai", 0) or 0), MAX_AI_IMAGES)
    system_prompt = row["system_prompt"] or DEFAULT_SYSTEM_PROMPT
    styl_flux     = _resolve_styl_flux(row["styl_flux"])
    historia_txt  = _format_historia(historia)

    # Python oblicza liczbę dni — AI nie liczy tego poprawnie
    dni = _oblicz_dni(data_smierci_str)
    system_prompt = (system_prompt
        .replace("{data_smierci_str}", data_smierci_str)
        .replace("{dni}", dni)
    )
    current_app.logger.info("[smierc] etap=%d dni_od_smierci=%s", etap, dni)

    # ── Tekst odpowiedzi ──────────────────────────────────────────────────────
    user_msg = (
        f"Etap: {opis}\n"
        f"Wiadomość od nadawcy: {body}\n"
        f"Historia:\n{historia_txt}"
    )
    wynik = call_deepseek(system_prompt, user_msg, MODEL_TYLER)
    if not wynik:
        current_app.logger.warning("[smierc] DeepSeek zawiódł etap=%d — próbuję Groq", etap)
        wynik = _call_groq(system_prompt, user_msg)

    reply_html = (
        f"<p>{wynik}</p>" if wynik
        else "<p>Chwilowo brak zasięgu w tej strefie kosmicznej.</p>"
    )

    # ── Obrazy statyczne ──────────────────────────────────────────────────────
    images = _load_images(obraz_lista) if obraz_lista else []
    if images:
        current_app.logger.info("[smierc] etap=%d obrazy statyczne: %d", etap, len(images))

    # ── Obrazki AI (FLUX) — generowane równolegle, każdy inny prompt ──────────
    images_ai = []
    debug_txt = None

    if obrazki_ai > 0:
        current_app.logger.info("[smierc] etap=%d obrazki_ai=%d START", etap, obrazki_ai)
        images_ai, debug_lines = _generate_n_flux_images(
            obrazki_ai, wynik or opis, styl_flux, etap
        )
        # Kompresuj PNG → JPG
        images_ai = _compress_images_ai(images_ai, obrazki_ai)
        debug_txt = _build_debug_txt(wynik or "", debug_lines, etap)
        current_app.logger.info("[smierc] etap=%d obrazki_ai=%d/%d udanych",
                                etap, len(images_ai), obrazki_ai)
    else:
        current_app.logger.info("[smierc] etap=%d obrazki_ai=0 — pomijam FLUX", etap)

    # ── Video ─────────────────────────────────────────────────────────────────
    videos = _load_videos(video_lista) if video_lista else []

    current_app.logger.info(
        "[smierc] etap=%d reply=%s images=%d images_ai=%d videos=%d",
        etap, bool(wynik), len(images), len(images_ai), len(videos)
    )

    return {
        "reply_html": reply_html,
        "nowy_etap":  etap + 1,
        "images":     images,
        "images_ai":  images_ai,
        "videos":     videos,
        "debug_txt":  debug_txt,
    }
